import pandas as pd
import numpy as np
import os
from openpyxl import load_workbook

dir_principal = '/Users/yurim/Documents/python/gera_relatorios'
#dir_principal = '/Users/Leonardo/Desktop/Tocalivros Dropbox/Campsoft/Geral/Relatórios/provedores'

### Verifica se o diretório existe e o cria se não existir. ###
def criar_diretorio(caminho):
    if not os.path.exists(caminho):
        os.makedirs(caminho)
        print(f"Diretório criado: {caminho}")
    else:
        print(f"O diretório já existe: {caminho}")


# CRIA RELATÓRIO BRUTO
def gera_latorio_bruto(dir_provedor, arquivo_nome, df):
    criar_diretorio(f"{dir_provedor}/bruto/")
    df.to_excel(f"{dir_provedor}/bruto/{arquivo_nome}", index=False)


# CRIA RELATÓRIO DETALHADO
def gera_latorio_detalhado(dir_provedor, arquivo_nome, df):
    # Seleciona apenas as colunas necessárias
    df_detalhado = df[['ano', 'mes', 'provedor_nome', 'conteudista_produto_nome', 'nome', 'cpf_cnpj', 'email', 'data_inicio', 'data_final']]
    criar_diretorio(f"{dir_provedor}/detalhado/")

    # Salva excel
    df_detalhado.to_excel(f"{dir_provedor}/detalhado/{arquivo_nome}", index=False)

    # Abre o arquivo Excel com openpyxl
    workbook = load_workbook(f"{dir_provedor}/detalhado/{arquivo_nome}")
    worksheet = workbook.active

    # Define o tamanho das colunas
    col_widths = {
        'A': 5,  # ano
        'B': 5,  # mes
        'C': 20,  # provedor_nome
        'D': 25,  # conteudista_produto_nome
        'E': 50,  # nome
        'F': 17,  # cpf_cnpj
        'G': 50,  # email
        'H': 18,  # data_inicio
        'I': 18  # data_inicio
    }

    # Ajusta o tamanho das colunas
    for col, width in col_widths.items():
        worksheet.column_dimensions[col].width = width

    # Salva o arquivo Excel com as colunas redimensionadas
    workbook.save(f"{dir_provedor}/detalhado/{arquivo_nome}")
    workbook.close()


# CRIA RELATÓRIO RESUMIDO
def gera_latorio_resumido(dir_provedor, arquivo_nome, df):
    df_resumido = (
        df.groupby(['ano', 'mes', 'conteudista_produto_nome'])
        .agg({'id_provedor_usuario': pd.Series.nunique})
        .reset_index()
        .rename(columns={'id_provedor_usuario': 'assinaturas'})
    )
    criar_diretorio(f"{dir_provedor}/resumido/")
    df_resumido.to_excel(f"{dir_provedor}/resumido/{arquivo_nome}", index=False)

    # Abre o arquivo Excel com openpyxl
    workbook = load_workbook(f"{dir_provedor}/resumido/{arquivo_nome}")
    worksheet = workbook.active

    # Define o tamanho das colunas
    col_widths = {
        'A': 5,  # ano
        'B': 5,  # mes
        'C': 26,  # conteudista_produto_nome
        'D': 10,  # assinaturas
    }

    # Ajusta o tamanho das colunas
    for col, width in col_widths.items():
        worksheet.column_dimensions[col].width = width

    # Salva o arquivo Excel com as colunas redimensionadas
    workbook.save(f"{dir_provedor}/resumido/{arquivo_nome}")
    workbook.close()


def gera_latorio(id_provedor, nome_provedor, df):
    try:
        # Verifica se o DataFrame está vazio
        if df.empty:
            print(f"DataFrame vazio para o provedor {id_provedor}. Nenhum arquivo será gerado.")
            return

        # Verifica se o diretório principal existe
        if not os.path.exists(dir_principal):
            print(f"Diretório principal não encontrado: {dir_principal}")
            return

        # Verifica se as colunas necessárias estão presentes
        colunas_necessarias = ['ano', 'mes', 'provedor_nome', 'conteudista_produto_nome', 'nome', 'cpf_cnpj', 'email', 'data_inicio', 'data_final', 'id_provedor_usuario']
        if not all(coluna in df.columns for coluna in colunas_necessarias):
            print(f"Colunas necessárias não encontradas no DataFrame para o provedor {id_provedor}.")
            return

        ano = df.iloc[0]['ano']
        mes = df.iloc[0]['mes']
        dir_provedor = f"{dir_principal}/{id_provedor}"
        arquivo_nome = f"{ano}.{mes}-{nome_provedor}.xlsx"

        print(f"Gerando relatório para: Ano={ano}, Mês={mes}, Provedor={nome_provedor}")

        gera_latorio_bruto(dir_provedor, arquivo_nome, df)
        gera_latorio_detalhado(dir_provedor, arquivo_nome, df)
        gera_latorio_resumido(dir_provedor, arquivo_nome, df)

    except Exception as e:
        print(f"Erro ao gerar relatório para o provedor {id_provedor}: {e}")